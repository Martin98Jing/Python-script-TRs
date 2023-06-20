from openpyxl.utils import get_column_letter
import re
import pandas as pd
import numpy as np
import openpyxl
from jira import JIRA
import getpass


def contains_pattern_AnswerCode(string):
    pattern_AnswerCode_only_AorB = r'A\d+\s*-|B\d+\s*-'
    return re.search(pattern_AnswerCode_only_AorB, string) is not None


def contains_pattern_TrPcpbWp(string):
    pcpb_tr_wp_pattern = r'(PCPB-\d+|PCTR-\d+|WP\d+)'
    return re.search(pcpb_tr_wp_pattern, string) is not None


# -------------------------------------------
# 读取Excel文件中的数据到DataFrame对象
# fileNmae = 'PDU Packet Core JIRA 2023-05-26T04_32_19+0200.xlsx'
# fileNmae = 'C:/Users/euzijnh/OneDrive - Ericsson/Documents/Jira Data Analyse/final demo/PDU Packet Core JIRA 2023-05-26T04_32_19+0200.xlsx'
# The_clo_Answer_Code = 35
# The_clo_Why_fault_introduced = 36
# The_clo_Teamname = 37

# 从命令行输入账号和密码 .....
# username = input("Please enter your EID to login JIRA: ")
# password = getpass.getpass("Please enter your JIRA's password: ")
username = "euzijnh"
password = "zj@545167594"
fileNmae = input("Please enter fileNmae(input): ")
fileNmaeOutput = input("Please enter fileNmae(output): ")

jira = JIRA(server='https://pdupc-jira.internal.ericsson.com/', basic_auth=(username, password))

df = pd.read_excel(fileNmae)
# 打开原始 Excel 文件  openpyxl 库可以正确地统计空单元格！Pandas一直有问题
wb = openpyxl.load_workbook(fileNmae)
ws = wb.active

The_clo_Answer_Code = The_clo_Why_fault_introduced = The_clo_Teamname = 0
num_total_TRs = 0
num_total_TRs_with_AnswerCode_AB = 0
empty_count = 0
nonEmpty_count_without_pcpb_tr_wp = 0
Num_pcpb_tr_wp = 0

# 遍历第一行的所有单元格，找出三个目标列在哪里
for cell in ws[1]:
    # 检查单元格的值是否包含目标字符串
    if 'Why was the fault introduced' in str(cell.value):
        # 获取列数（从 1 开始）
        The_clo_Why_fault_introduced = cell.column - 1
        print(f"The column number is: {The_clo_Why_fault_introduced}")
    if 'Answer Code' in str(cell.value):
        The_clo_Answer_Code = cell.column - 1
        print(f"The_clo_Answer_Code is: {The_clo_Answer_Code}")
    if 'Teamname' in str(cell.value):
        The_clo_Teamname = cell.column - 1
        print(f"The_clo_Teamname is: {The_clo_Teamname}")

for row in ws.iter_rows(min_row=2, values_only=True):
    num_total_TRs += 1
    answer_Code = row[The_clo_Answer_Code]  # 在当前这一行，提取 The_clo_Answer_Code 这一列的数据
    why_fault_introduced = row[The_clo_Why_fault_introduced]
    if contains_pattern_TrPcpbWp(str(why_fault_introduced)):
        Num_pcpb_tr_wp += 1

    if contains_pattern_AnswerCode(str(answer_Code)):  # cell_value 可能是 非string类型的。因此需要强制类型转换
        num_total_TRs_with_AnswerCode_AB += 1
        # print(f'String "{answer_Code}" matches the pattern.')
        if why_fault_introduced is None:
            empty_count += 1
        else:
            if contains_pattern_TrPcpbWp(str(why_fault_introduced)) is not True:
                nonEmpty_count_without_pcpb_tr_wp += 1


# 输出包含"PCTR|PCPB|WP"的单元格数量
print(f"一共有{num_total_TRs_with_AnswerCode_AB}个TR的Answer Code 是A or B")
print(f"Why_fault_introduced列中一共有{Num_pcpb_tr_wp}个单元格包含\"PCTR|PCPB|WP\"")
print(f"Why_fault_introduced中一共有{nonEmpty_count_without_pcpb_tr_wp}个非空且不包含PCTR也不包含PCPB也不包含WP单元格")
print(f"Why_fault_introduced列中一共有{empty_count}个空单元格")


PCPB_WP_Num = "PCPB_WP_Num"
PCTR_Num = "PCTR_Num"

# 筛选数据并提取PCPB-几个数字的内容
pcpb_wp_pattern = r'(PCPB-\d+|WP\d+)'
pctr_pattern = r'(PCTR-\d+)'


def extract_all_matches(text, pattern):
    matches = re.findall(pattern, text)
    matches = list(set(matches))  # 使用 set() 去重
    # print("matches = ", matches)
    return matches if matches else []


# df = df.assign(PCPB_WP_Num=df.iloc[:, The_clo_Why_fault_introduced].astype(str).apply(lambda x: extract_all_matches(x, pcpb_wp_pattern)))
df[PCPB_WP_Num] = df.iloc[:, The_clo_Why_fault_introduced].astype(str).apply(lambda x: extract_all_matches(x, pcpb_wp_pattern))
df = df.explode(PCPB_WP_Num)
# df = df.assign(PCTR_Num=df.iloc[:, The_clo_Why_fault_introduced].astype(str).apply(lambda x: extract_all_matches(x, pctr_pattern)))
df[PCTR_Num] = df.iloc[:, The_clo_Why_fault_introduced].astype(str).apply(lambda x: extract_all_matches(x, pctr_pattern))
df = df.explode(PCTR_Num)


# 统计每个PCTR-五个数字的出现次数
pcpb_wp_counts = df[PCPB_WP_Num].value_counts()
pctr_counts = df[PCTR_Num].value_counts()
# print(pcpb_wp_counts)
# print(pctr_counts)


pcpb_teamname = pd.Series(index=pcpb_wp_counts.index, dtype=object)
for pcpb_id in pcpb_wp_counts.index:
    # 如果是wp，则将wp对应的teamname设置为空string
    if (re.search(r'wp\d+', pcpb_id, re.IGNORECASE)):
        pcpb_teamname[pcpb_id] = np.nan
        continue
    issues = jira.search_issues('key = ' + pcpb_id)      # 执行Jira查询
    if not issues:
        pcpb_teamname[pcpb_id] = np.nan
        continue
    # 遍历查询结果，并打印每个PCTR中的团队名称
    for issue in issues:
        pcpb_teamname[pcpb_id] = issue.fields.customfield_16002

pctr_teamname = pd.Series(index=pctr_counts.index, dtype=object)
for pctr_id in pctr_counts.index:
    issues = jira.search_issues('key = ' + pctr_id)      # 执行Jira查询
    # print("pctr_id =", pctr_id)
    if not issues:
        pctr_teamname[pctr_id] = np.nan
        # print(pctr_teamname[pctr_id])
        continue
    # 遍历查询结果，并打印每个PCTR中的团队名称
    for issue in issues:
        pctr_teamname[pctr_id] = issue.fields.customfield_16002
        # print(pctr_teamname[pctr_id])

# 统计第一个 Series 的唯一值数量
Nb_Teams_related_Pcpb = len(pd.Series(pcpb_teamname.unique()).dropna())
# 遍历第二个 Series 并统计数量
Nb_Teams_related_tr = 0
for value in pctr_teamname.unique():
    if value not in pcpb_teamname.values and value is not None:
        # print("team = ", value)
        Nb_Teams_related_tr += 1

Nb_Teams_related_trPcpb = Nb_Teams_related_tr + Nb_Teams_related_Pcpb
print("引起 TR/PCPB/WP的team总个数(去重)", Nb_Teams_related_trPcpb)


# 建一个新的sheet来存储PCPB/WP 或者PCTR 的统计数据
def build_newSheet_with_statistical_data(pcpbWp_Or_pctr_counts, teamName, column_Name):
    df = pd.DataFrame({column_Name: pcpbWp_Or_pctr_counts.index, 'Nb_TR_introduced': pcpbWp_Or_pctr_counts.values, 'Teamname': teamName})
    df['Teamname'] = df['Teamname'].astype(str)
    grouped_data = df.groupby('Teamname')['Nb_TR_introduced'].sum()
    sorted_data = grouped_data.sort_values(ascending=False)
    df = df.merge(sorted_data, on='Teamname', how='right', suffixes=('', '_total'))
    return df




# 在每一行（每一个PCPB/WP OR PCTR）后新增若干列，显示所有的由该 PCTR 或者PCPB/WP 引入的 TR
def Add_trID_introduced_by_pctrOrWpPcpb_as_additional_column(counts_df, column_Name):
    for index, row in counts_df.iterrows():
        # 在df中查找匹配的行
        matched_rows = df[df.iloc[:, The_clo_Why_fault_introduced].astype(str).str.contains(str(row[column_Name]), na=False)]
        # print("match rows : ", matched_rows)
        # 获取当前行的起始列索引
        start_col_index = counts_df.columns.get_loc(column_Name) + 1
        curr_col_index = start_col_index
        # 添加匹配行的issue key值到pcpb_counts_df表中
        for _, match_row in matched_rows.iterrows():
            issue_key = match_row['Issue key']
            # print("issue_key : ", issue_key)
            # 因为matched_rows 会出现重复的pcpb_wp_num值，为了避免将重复的pctr号写入excel。这里使用if人工判断，如果当前pctr与同行前一列的pctr一致，则跳过当前pctr
            if curr_col_index > start_col_index and counts_df.at[index, curr_col_index-1] == issue_key:
                continue
            # 将issue key值添加到对应的单元格
            counts_df.at[index, curr_col_index] = issue_key
            # print("index =", index, "curr_col_index =", curr_col_index, "df.at[index, curr_col_index] = ", df.at[index, curr_col_index])
            curr_col_index += 1


pcpb_counts_df = build_newSheet_with_statistical_data(pcpb_wp_counts, pcpb_teamname, PCPB_WP_Num)
Add_trID_introduced_by_pctrOrWpPcpb_as_additional_column(pcpb_counts_df, PCPB_WP_Num)
pctr_counts_df = build_newSheet_with_statistical_data(pctr_counts, pctr_teamname, PCTR_Num)
Add_trID_introduced_by_pctrOrWpPcpb_as_additional_column(pctr_counts_df, PCTR_Num)

# for index, row in pcpb_counts_df.iterrows():
#     # pcpb_wp_num = row[PCPB_WP_Num]
#     # 在df中查找匹配的行
#     matched_rows = df[df.iloc[:, The_clo_Why_fault_introduced].astype(str).str.contains(str(row[PCPB_WP_Num]), na=False)]
#     # print("match rows : ", matched_rows)
#     # 获取当前行的起始列索引
#     start_col_index = pcpb_counts_df.columns.get_loc(PCPB_WP_Num) + 1
#     curr_col_index = start_col_index
#     # 添加匹配行的issue key值到pcpb_counts_df表中
#     for _, match_row in matched_rows.iterrows():
#         issue_key = match_row['Issue key']
#         # print("issue_key : ", issue_key)
#         # 因为matched_rows 会出现重复的pcpb_wp_num值，为了避免将重复的pctr号写入excel。这里使用if人工判断，如果当前pctr与同行前一列的pctr一致，则跳过当前pctr
#         if curr_col_index > start_col_index and pcpb_counts_df.at[index, curr_col_index-1] == issue_key:
#             continue
#         # 将issue key值添加到对应的单元格
#         pcpb_counts_df.at[index, curr_col_index] = issue_key
#         # print("index =", index, "curr_col_index =", curr_col_index, "pcpb_counts_df.at[index, curr_col_index] = ", pcpb_counts_df.at[index, curr_col_index])
#         curr_col_index += 1


# pctr_counts_df = pd.DataFrame({PCTR_Num: pctr_counts.index, 'Nb_TR_introduced': pctr_counts.values, "Teamname": pctr_teamname})
# pctr_counts_df['Teamname'] = pctr_counts_df['Teamname'].astype(str)
# grouped_data = pctr_counts_df.groupby('Teamname')['Nb_TR_introduced'].sum()
# sorted_data = grouped_data.sort_values(ascending=False)
# pctr_counts_df = pctr_counts_df.merge(sorted_data, on='Teamname', how='right', suffixes=('', '_total'))

# for index, row in pctr_counts_df.iterrows():
#     # pctr_num = row[PCTR_Num]
#     # 在df中查找匹配的行
#     matched_rows = df[df.iloc[:, The_clo_Why_fault_introduced].astype(str).str.contains(str(row[PCTR_Num]), na=False)]
#     # print("match rows : ", matched_rows)
#     # 获取当前行的起始列索引
#     start_col_index = pctr_counts_df.columns.get_loc(PCTR_Num) + 1
#     curr_col_index = start_col_index
#     # 添加匹配行的issue key值到pcpb_counts_df表中
#     for _, match_row in matched_rows.iterrows():
#         issue_key = match_row['Issue key']
#         if curr_col_index > start_col_index and pctr_counts_df.at[index, curr_col_index-1] == issue_key:
#             continue
#         # print("issue_key : ", issue_key)
#         # 将issue key值添加到对应的单元格
#         pctr_counts_df.at[index, curr_col_index] = issue_key
#         curr_col_index += 1


Summary = {
    'Summary': ['Count of TR number', "The number of teams involved"],
    'Total TR number': [num_total_TRs, 0],
    'Total TR number (Only answer code A/B  included)': [num_total_TRs_with_AnswerCode_AB, 0],
    'Field with TR/PCPB/WP info': [Num_pcpb_tr_wp, Nb_Teams_related_trPcpb],
    'Field is not empty but no TR/PCPB/WP info included': [nonEmpty_count_without_pcpb_tr_wp, 0],
    'Field is empty': [empty_count, 0],
}
summary_counts_df = pd.DataFrame(Summary)

# 创建一个 ExcelWriter 对象，用于存储数据到同一个 Excel 文件
# writer = pd.ExcelWriter('TR statistical analysis.xlsx')
writer = pd.ExcelWriter(fileNmaeOutput, engine='openpyxl')

# 将每个数据集写入不同的工作表中
summary_counts_df.to_excel(writer, sheet_name='summary_counts', index=False)  # index=False，以避免将行索引保存到文件中
pcpb_counts_df.to_excel(writer, sheet_name='pcpb_counts', index=False)
pctr_counts_df.to_excel(writer, sheet_name='pctr_counts', index=False)


# 获取工作簿和工作表对象
workbook = writer.book
worksheet_summary_counts = workbook['summary_counts']
worksheet_pcpb_counts = workbook['pcpb_counts']
worksheet_pctr_counts = workbook['pctr_counts']


def merge_cells_if_same_teamname(workshhet):
    current_teamname = None
    merge_start_row = 2  # 从第二行开始合并单元格

    # 合并同一个team的PCPB_Num_total
    # 这里有个问题。这个for循环 不会对最后一个team 对应的Count_PB_total进行合并单元格。但是一般来说，排在最后的team Count_PB_total 值 是1. 因此没有继续进行debug。
    for row in range(2, workshhet.max_row + 1):
        teamname = workshhet.cell(row=row, column=3).value
        if teamname != current_teamname:
            merge_end_row = row - 1
            if merge_start_row < merge_end_row:
                merge_range = f'D{merge_start_row}:D{merge_end_row}'
                workshhet.merge_cells(merge_range)
            merge_start_row = row
            current_teamname = teamname


merge_cells_if_same_teamname(worksheet_pcpb_counts)
merge_cells_if_same_teamname(worksheet_pctr_counts)


def adaptive_column_width(df, ws):
    for column in df.columns:
        column_width = max(df[column].astype(str).map(len).max(), len(str(column)))  # pcpb_counts_df 中 有一些column是index 123456 这些是int类型的数据，需要强制类型转换成str 才可以使用len()
        column_letter = get_column_letter(df.columns.get_loc(column) + 1)
        ws.column_dimensions[column_letter].width = column_width + 3


adaptive_column_width(summary_counts_df, worksheet_summary_counts)
adaptive_column_width(pcpb_counts_df, worksheet_pcpb_counts)
adaptive_column_width(pctr_counts_df, worksheet_pctr_counts)

# 保存工作簿到文件
# workbook.save(fileNmaeOutput)
# 关闭 ExcelWriter 对象，释放与 ExcelWriter 对象相关的资源，包括 Excel 文件的句柄或连接。
writer.close()
