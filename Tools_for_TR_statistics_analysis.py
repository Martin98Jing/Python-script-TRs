import re
from openpyxl.utils import get_column_letter
import pandas as pd
import numpy as np


def contains_pattern_AnswerCode(string):
    pattern_AnswerCode_only_AorB = r'A\d+\s*-|B\d+\s*-'
    return re.search(pattern_AnswerCode_only_AorB, string) is not None


def contains_pattern_TrPcpbWp(string):
    pcpb_tr_wp_pattern = r'(PCPB-\d+|PCTR-\d+|WP\d+)'
    return re.search(pcpb_tr_wp_pattern, string) is not None


def extract_all_matches(text, pattern):
    matches = re.findall(pattern, text)
    matches = list(set(matches))  # 使用 set() 去重
    # print("matches = ", matches)
    return matches if matches else []


def merge_cells_if_same_teamname(workshhet):
    current_teamname = None
    merge_start_row = 2  # 从第二行开始合并单元格

    # 合并同一个team的PCPB_Num_total
    # 这里有个问题。这个for循环 不会对最后一个team 对应的Count_PB_total进行合并单元格。但是一般来说，排在最后的team Count_PB_total 值 是1. 因此没有继续进行debug。
    for row in range(2, workshhet.max_row + 1):
        teamname = workshhet.cell(row=row, column=3).value
        if teamname != current_teamname:
            merge_end_row = row - 1
            if merge_start_row < merge_end_row:
                merge_range = f'D{merge_start_row}:D{merge_end_row}'
                workshhet.merge_cells(merge_range)
            merge_start_row = row
            current_teamname = teamname


def adaptive_column_width(df, ws):
    for column in df.columns:
        column_width = max(df[column].astype(str).map(len).max(), len(str(column)))  # pcpb_counts_df 中 有一些column是index 123456 这些是int类型的数据，需要强制类型转换成str 才可以使用len()
        column_letter = get_column_letter(df.columns.get_loc(column) + 1)
        ws.column_dimensions[column_letter].width = column_width + 3


# 构建teamname这一列Series,根据每个PCPB/PCTR id ，从jira上拉取teamname数据。通过抓包了解到，jira上返回的customfield_16002字段是teamname
def create_teamName_corresponding_pctr_or_pcpcWp(pcpbWp_Or_pctr_counts):
    teamName = pd.Series(index=pcpbWp_Or_pctr_counts.index, dtype=object)
    for pc_id in pcpbWp_Or_pctr_counts.index:
        # 如果是wp，则将wp对应的teamname设置为空string，因为在jira上一定找不到wp对应的teamname
        if (re.search(r'wp\d+', pc_id, re.IGNORECASE)):
            teamName[pc_id] = np.nan
            continue
        from TR_statistics_analysis import jira
        issues = jira.search_issues('key = ' + pc_id)      # 执行Jira查询
        if not issues:
            teamName[pc_id] = np.nan
            continue
        # 遍历查询结果，并打印每个PCTR中的团队名称
        for issue in issues:
            teamName[pc_id] = issue.fields.customfield_16002
    return teamName


# 建一个新的sheet来存储PCPB/WP 或者PCTR 的统计数据
def build_newSheet_with_statistical_data(pcpbWp_Or_pctr_counts, teamName, column_Name):
    df = pd.DataFrame({column_Name: pcpbWp_Or_pctr_counts.index, 'Nb_TR_introduced': pcpbWp_Or_pctr_counts.values, 'Teamname': teamName})
    df['Teamname'] = df['Teamname'].astype(str)
    grouped_data = df.groupby('Teamname')['Nb_TR_introduced'].sum()
    sorted_data = grouped_data.sort_values(ascending=False)
    df = df.merge(sorted_data, on='Teamname', how='right', suffixes=('', '_total'))  # 根据teamname计算总TR引入数量
    return df


# 在每一行（每一个PCPB/WP OR PCTR）后新增若干列，显示所有的由该 PCTR 或者PCPB/WP 引入的 TR
def Add_trID_introduced_by_pctrOrWpPcpb_as_additional_column(counts_df, column_Name):
    for index, row in counts_df.iterrows():
        # 在df中查找匹配的行
        from TR_statistics_analysis import The_clo_Why_fault_introduced, df
        matched_rows = df[df.iloc[:, The_clo_Why_fault_introduced].astype(str).str.contains(str(row[column_Name]), na=False)]
        # print("match rows : ", matched_rows)
        # 获取当前行的起始列索引
        start_col_index = counts_df.columns.get_loc(column_Name) + 1
        curr_col_index = start_col_index
        # 添加匹配行的issue key值到pcpb_counts_df表中
        for _, match_row in matched_rows.iterrows():
            issue_key = match_row['Issue key']
            # print("issue_key : ", issue_key)
            # 因为matched_rows 会出现重复的pcpb_wp_num值，为了避免将重复的pctr号写入excel。这里使用if人工判断，如果当前pctr与同行前一列的pctr一致，则跳过当前pctr
            if curr_col_index > start_col_index and counts_df.at[index, curr_col_index-1] == issue_key:
                continue
            # 将issue key值添加到对应的单元格
            counts_df.at[index, curr_col_index] = issue_key
            # print("index =", index, "curr_col_index =", curr_col_index, "df.at[index, curr_col_index] = ", df.at[index, curr_col_index])
            curr_col_index += 1