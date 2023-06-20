import pandas as pd
import openpyxl
from jira import JIRA
from Tools_for_TR_statistics_analysis import contains_pattern_AnswerCode, contains_pattern_TrPcpbWp, extract_all_matches, adaptive_column_width
from Tools_for_TR_statistics_analysis import merge_cells_if_same_teamname, create_teamName_corresponding_pctr_or_pcpcWp, build_newSheet_with_statistical_data, Add_trID_introduced_by_pctrOrWpPcpb_as_additional_column
import getpass

# -------------------------------------------
# 读取Excel文件中的数据到DataFrame对象
# fileNmae = 'PDU Packet Core JIRA 2023-05-26T04_32_19+0200.xlsx'  # 相对路径要求excel和执行的python文件在同一目录下，以及当前已经处于此目录
# fileNmae = 'C:/Users/euzijnh/OneDrive - Ericsson/Documents/Jira Data Analyse/final demo/PDU Packet Core JIRA 2023-05-26T04_32_19+0200.xlsx'  # 绝对路径

# 从命令行输入账号和密码 .....
username = input("Please enter your EID to login JIRA: ")
password = getpass.getpass("Please enter your JIRA's password: ")
fileNmae = input("Please enter fileNmae(input): ")
fileNmaeOutput = input("Please enter fileNmae(output): ")

jira = JIRA(server='https://pdupc-jira.internal.ericsson.com/', basic_auth=(username, password))
df = pd.read_excel(fileNmae)
# 打开原始 Excel 文件  openpyxl 库可以正确地统计空单元格！Pandas一直有问题
wb = openpyxl.load_workbook(fileNmae)
ws = wb.active

The_clo_Answer_Code = The_clo_Why_fault_introduced = The_clo_Teamname = 0
num_total_TRs = 0
num_total_TRs_with_AnswerCode_AB = 0
empty_count = 0
nonEmpty_count_without_pcpb_tr_wp = 0
Num_pcpb_tr_wp = 0

# 遍历第一行的所有单元格，找出三个目标列在哪里
for cell in ws[1]:
    # 检查单元格的值是否包含目标字符串
    if 'Why was the fault introduced' in str(cell.value):
        # 获取列数（从 1 开始）
        The_clo_Why_fault_introduced = cell.column - 1
        print(f"The column number is: {The_clo_Why_fault_introduced}")
    if 'Answer Code' in str(cell.value):
        The_clo_Answer_Code = cell.column - 1
        print(f"The_clo_Answer_Code is: {The_clo_Answer_Code}")
    if 'Teamname' in str(cell.value):
        The_clo_Teamname = cell.column - 1
        print(f"The_clo_Teamname is: {The_clo_Teamname}")

for row in ws.iter_rows(min_row=2, values_only=True):
    num_total_TRs += 1
    answer_Code = row[The_clo_Answer_Code]  # 在当前这一行，提取 The_clo_Answer_Code 这一列的数据
    why_fault_introduced = row[The_clo_Why_fault_introduced]
    if contains_pattern_TrPcpbWp(str(why_fault_introduced)):
        Num_pcpb_tr_wp += 1

    if contains_pattern_AnswerCode(str(answer_Code)):  # cell_value 可能是 非string类型的。因此需要强制类型转换
        num_total_TRs_with_AnswerCode_AB += 1
        # print(f'String "{answer_Code}" matches the pattern.')
        if why_fault_introduced is None:
            empty_count += 1
        else:
            if contains_pattern_TrPcpbWp(str(why_fault_introduced)) is not True:
                nonEmpty_count_without_pcpb_tr_wp += 1

print(f"一共有{num_total_TRs_with_AnswerCode_AB}个TR的Answer Code 是A or B")
print(f"Why_fault_introduced列中一共有{Num_pcpb_tr_wp}个单元格包含\"PCTR|PCPB|WP\"")
print(f"Why_fault_introduced中一共有{nonEmpty_count_without_pcpb_tr_wp}个非空且不包含PCTR也不包含PCPB也不包含WP单元格")
print(f"Why_fault_introduced列中一共有{empty_count}个空单元格")


PCPB_WP_Num = "PCPB_WP_Num"
PCTR_Num = "PCTR_Num"
# 筛选数据并提取PCPB-几个数字的内容
pcpb_wp_pattern = r'(PCPB-\d+|WP\d+)'
pctr_pattern = r'(PCTR-\d+)'

# df = df.assign(PCPB_WP_Num=df.iloc[:, The_clo_Why_fault_introduced].astype(str).apply(lambda x: extract_all_matches(x, pcpb_wp_pattern)))
df[PCPB_WP_Num] = df.iloc[:, The_clo_Why_fault_introduced].astype(str).apply(lambda x: extract_all_matches(x, pcpb_wp_pattern))
df = df.explode(PCPB_WP_Num)
# df = df.assign(PCTR_Num=df.iloc[:, The_clo_Why_fault_introduced].astype(str).apply(lambda x: extract_all_matches(x, pctr_pattern)))
df[PCTR_Num] = df.iloc[:, The_clo_Why_fault_introduced].astype(str).apply(lambda x: extract_all_matches(x, pctr_pattern))
df = df.explode(PCTR_Num)

# 统计每个PCTR-五个数字的出现次数
pcpb_wp_counts = df[PCPB_WP_Num].value_counts()
pctr_counts = df[PCTR_Num].value_counts()

# 构建teamname这一列Series
pcpb_teamname = create_teamName_corresponding_pctr_or_pcpcWp(pcpb_wp_counts)
pctr_teamname = create_teamName_corresponding_pctr_or_pcpcWp(pctr_counts)

# 统计第一个 Series 的唯一值数量
Nb_Teams_related_Pcpb = len(pd.Series(pcpb_teamname.unique()).dropna())
# 遍历第二个 Series 并统计数量
Nb_Teams_related_tr = 0
for value in pctr_teamname.unique():
    if value not in pcpb_teamname.values and value is not None:
        # print("team = ", value)
        Nb_Teams_related_tr += 1

Nb_Teams_related_trPcpb = Nb_Teams_related_tr + Nb_Teams_related_Pcpb
print("引起 TR/PCPB/WP的team总个数(去重)", Nb_Teams_related_trPcpb)

# 建一个新的sheet来存储PCPB/WP 或者PCTR 的统计数据
pcpb_counts_df = build_newSheet_with_statistical_data(pcpb_wp_counts, pcpb_teamname, PCPB_WP_Num)
Add_trID_introduced_by_pctrOrWpPcpb_as_additional_column(pcpb_counts_df, PCPB_WP_Num)
pctr_counts_df = build_newSheet_with_statistical_data(pctr_counts, pctr_teamname, PCTR_Num)
Add_trID_introduced_by_pctrOrWpPcpb_as_additional_column(pctr_counts_df, PCTR_Num)

Summary = {
    'Summary': ['Count of TR number', "The number of teams involved"],
    'Total TR number': [num_total_TRs, 0],
    'Total TR number (Only answer code A/B  included)': [num_total_TRs_with_AnswerCode_AB, 0],
    'Field with TR/PCPB/WP info': [Num_pcpb_tr_wp, Nb_Teams_related_trPcpb],
    'Field is not empty but no TR/PCPB/WP info included': [nonEmpty_count_without_pcpb_tr_wp, 0],
    'Field is empty': [empty_count, 0],
}
summary_counts_df = pd.DataFrame(Summary)

# 创建一个 ExcelWriter 对象，用于存储数据到同一个 Excel 文件
writer = pd.ExcelWriter(fileNmaeOutput, engine='openpyxl')

# 将每个数据集写入不同的工作表中
summary_counts_df.to_excel(writer, sheet_name='summary_counts', index=False)  # index=False，以避免将行索引保存到文件中
pcpb_counts_df.to_excel(writer, sheet_name='pcpb_counts', index=False)
pctr_counts_df.to_excel(writer, sheet_name='pctr_counts', index=False)

# 获取工作簿和工作表对象
workbook = writer.book
worksheet_summary_counts = workbook['summary_counts']
worksheet_pcpb_counts = workbook['pcpb_counts']
worksheet_pctr_counts = workbook['pctr_counts']

merge_cells_if_same_teamname(worksheet_pcpb_counts)
merge_cells_if_same_teamname(worksheet_pctr_counts)

adaptive_column_width(summary_counts_df, worksheet_summary_counts)
adaptive_column_width(pcpb_counts_df, worksheet_pcpb_counts)
adaptive_column_width(pctr_counts_df, worksheet_pctr_counts)

# 关闭 ExcelWriter 对象，释放与 ExcelWriter 对象相关的资源，包括 Excel 文件的句柄或连接。
writer.close()
